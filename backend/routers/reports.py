from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from database import get_db
from models import Product, StockLog, User, Course
from auth import get_current_user, SECRET_KEY, ALGORITHM
from jose import jwt
from typing import Optional, List
from datetime import datetime, timezone, timedelta

TZ_TR = timezone(timedelta(hours=3))

def to_tr_time(dt):
    """UTC datetime'ı Türkiye saatine (UTC+3) çevirir."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(TZ_TR).isoformat()

import io

router = APIRouter(prefix="/api/reports", tags=["Reports"])

def get_optional_query_user(
    token: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    if not token:
        raise HTTPException(status_code=401, detail="Token bulunumadı")
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401)
    except Exception:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(user_id)).first()
    if not user:
        raise HTTPException(status_code=401)
    return user


@router.get("/logs")
def get_all_logs(
    product_id: Optional[int] = None,
    user_id: Optional[int] = None,
    action: Optional[str] = None,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(StockLog).options(
        joinedload(StockLog.user),
        joinedload(StockLog.product)
    )

    if product_id:
        query = query.filter(StockLog.product_id == product_id)
    if user_id:
        query = query.filter(StockLog.user_id == user_id)
    if action:
        query = query.filter(StockLog.action == action)

    # Filter by user's warehouse access
    if current_user.role != "admin":
        accessible_warehouse_ids = [w.id for w in current_user.warehouses]
        from sqlalchemy import or_
        if accessible_warehouse_ids:
            query = query.join(StockLog.product).filter(
                or_(
                    Product.warehouse_id.in_(accessible_warehouse_ids),
                    Product.warehouse_id.is_(None)
                )
            )
        else:
            query = query.join(StockLog.product).filter(Product.warehouse_id.is_(None))

    logs = query.order_by(StockLog.created_at.desc()).limit(limit).all()

    return [{
        "id": log.id,
        "product_id": log.product_id,
        "product_name": log.product.name if log.product else None,
        "user_id": log.user_id,
        "user_name": log.user.full_name if log.user else None,
        "action": log.action,
        "quantity_change": log.quantity_change,
        "previous_value": log.previous_value,
        "new_value": log.new_value,
        "note": log.note,
        "created_at": to_tr_time(log.created_at)
    } for log in logs]


@router.get("/export/excel")
def export_excel(
    category: Optional[str] = None,
    warehouse_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_optional_query_user)
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = db.query(Product).options(
        joinedload(Product.courses),
        joinedload(Product.warehouse)
    )

    if category and category != "Tümü":
        query = query.filter(Product.category == category)
    if warehouse_id:
        query = query.filter(Product.warehouse_id == warehouse_id)
    if status:
        query = query.filter(Product.status == status)

    # Filter by user's warehouse access
    if current_user.role != "admin":
        accessible_warehouse_ids = [w.id for w in current_user.warehouses]
        from sqlalchemy import or_
        if accessible_warehouse_ids:
            query = query.filter(
                or_(
                    Product.warehouse_id.in_(accessible_warehouse_ids),
                    Product.warehouse_id.is_(None)
                )
            )
        else:
            query = query.filter(Product.warehouse_id.is_(None))

    products = query.all()

    # Deduplicate
    seen = set()
    unique_products = []
    for p in products:
        if p.id not in seen:
            seen.add(p.id)
            unique_products.append(p)
    products = unique_products

    wb = Workbook()
    ws = wb.active
    ws.title = "Ürün Raporu"

    # Header styling
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        "No", "Ürün Adı", "Açıklama", "Kategori", "Dersler",
        "Depo", "Toplam Stok", "Kritik Stok", "İdeal Stok",
        "Durum", "Son Güncelleme"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row, product in enumerate(products, 2):
        courses_str = ", ".join([c.name for c in product.courses]) if product.courses else ""
        warehouse_name = product.warehouse.name if product.warehouse else ""

        values = [
            row - 1,
            product.name,
            product.description or "",
            product.category,
            courses_str,
            warehouse_name,
            product.current_stock,
            product.critical_stock,
            product.ideal_stock,
            product.status,
            product.updated_at.replace(tzinfo=timezone.utc).astimezone(TZ_TR).strftime("%d.%m.%Y %H:%M") if product.updated_at else ""
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Auto-width
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(int(max_length) + 2, 50)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"deneyap_stok_raporu_{datetime.now(timezone.utc).astimezone(TZ_TR).strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/pdf")
def export_pdf(
    category: Optional[str] = None,
    warehouse_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_optional_query_user)
):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    query = db.query(Product).options(
        joinedload(Product.courses),
        joinedload(Product.warehouse)
    )

    if category and category != "Tümü":
        query = query.filter(Product.category == category)
    if warehouse_id:
        query = query.filter(Product.warehouse_id == warehouse_id)
    if status:
        query = query.filter(Product.status == status)

    # Filter by user's warehouse access
    if current_user.role != "admin":
        accessible_warehouse_ids = [w.id for w in current_user.warehouses]
        from sqlalchemy import or_
        if accessible_warehouse_ids:
            query = query.filter(
                or_(
                    Product.warehouse_id.in_(accessible_warehouse_ids),
                    Product.warehouse_id.is_(None)
                )
            )
        else:
            query = query.filter(Product.warehouse_id.is_(None))

    products = query.all()
    seen = set()
    unique_products = []
    for p in products:
        if p.id not in seen:
            seen.add(p.id)
            unique_products.append(p)
    products = unique_products

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=20*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()

    elements = []
    title = Paragraph("Deneyap Atölyesi - Stok Raporu", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 10*mm))

    date_text = Paragraph(f"Rapor Tarihi: {datetime.now(timezone.utc).astimezone(TZ_TR).strftime('%d.%m.%Y %H:%M')}", styles['Normal'])
    elements.append(date_text)
    elements.append(Spacer(1, 5*mm))

    data = [["No", "Ürün Adı", "Kategori", "Stok", "Kritik", "Durum", "Depo"]]

    for i, product in enumerate(products, 1):
        warehouse_name = product.warehouse.name if product.warehouse else "-"
        data.append([
            str(i),
            product.name[:30],
            product.category[:20],
            str(product.current_stock),
            str(product.critical_stock),
            product.status[:15],
            warehouse_name[:20]
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F4F8')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    elements.append(table)
    doc.build(elements)
    output.seek(0)

    filename = f"deneyap_stok_raporu_{datetime.now(timezone.utc).astimezone(TZ_TR).strftime('%Y%m%d_%H%M')}.pdf"

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
