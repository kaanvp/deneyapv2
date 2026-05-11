from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Form
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_, func
from database import get_db
from datetime import timezone
from models import Product, Course, Warehouse, Cabinet, Shelf, Compartment, StockLog, product_courses
from schemas import ProductCreate, ProductUpdate, ProductResponse, CourseResponse, StockUpdate, StockLogCreate
from auth import get_current_user, get_admin_user
from models import User
from typing import List, Optional
import os
import uuid
import shutil
from supabase_utils import upload_file_to_supabase


router = APIRouter(prefix="/api/products", tags=["Products"])

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)


def product_to_response(product: Product) -> ProductResponse:
    return ProductResponse(
        id=product.id,
        name=product.name,
        description=product.description,
        image_url=product.image_url,
        category=product.category,
        status=product.status,
        current_stock=product.current_stock,
        critical_stock=product.critical_stock,
        ideal_stock=product.ideal_stock,
        warehouse_id=product.warehouse_id,
        cabinet_id=product.cabinet_id,
        shelf_id=product.shelf_id,
        compartment_id=product.compartment_id,
        created_at=product.created_at,
        updated_at=product.updated_at,
        courses=[CourseResponse(id=c.id, name=c.name) for c in product.courses],
        warehouse_name=product.warehouse.name if product.warehouse else None,
        cabinet_name=product.cabinet.name if product.cabinet else None,
        shelf_name=product.shelf.name if product.shelf else None,
        compartment_name=product.compartment.name if product.compartment else None,
    )


@router.get("/", response_model=List[ProductResponse])
def list_products(
    search: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[str] = None,
    warehouse_id: Optional[int] = None,
    course_id: Optional[int] = None,
    low_stock: Optional[bool] = None,
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(Product).options(
        joinedload(Product.courses),
        joinedload(Product.warehouse),
        joinedload(Product.cabinet),
        joinedload(Product.shelf),
        joinedload(Product.compartment)
    )

    # Filter by user's warehouse access (admin sees all)
    if current_user.role != "admin":
        accessible_warehouse_ids = [w.id for w in current_user.warehouses]
        if accessible_warehouse_ids:
            query = query.filter(
                or_(
                    Product.warehouse_id.in_(accessible_warehouse_ids),
                    Product.warehouse_id.is_(None)
                )
            )
        else:
            query = query.filter(Product.warehouse_id.is_(None))

    if search:
        search_term = f"%{search.lower()}%"
        query = query.filter(
            or_(
                func.lower(Product.name).like(search_term),
                func.lower(Product.description).like(search_term)
            )
        )

    if category and category != "Tümü":
        query = query.filter(Product.category == category)

    if status:
        query = query.filter(Product.status == status)

    if warehouse_id:
        query = query.filter(Product.warehouse_id == warehouse_id)

    if course_id:
        query = query.filter(Product.courses.any(Course.id == course_id))

    if low_stock:
        query = query.filter(Product.current_stock <= Product.critical_stock)

    total = query.count()
    products = query.order_by(Product.id.desc()).offset((page - 1) * per_page).limit(per_page).all()

    # Deduplicate products (joinedload may cause duplicates)
    seen = set()
    unique_products = []
    for p in products:
        if p.id not in seen:
            seen.add(p.id)
            unique_products.append(p)

    return [product_to_response(p) for p in unique_products]


@router.get("/count")
def get_product_count(
    search: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[str] = None,
    warehouse_id: Optional[int] = None,
    low_stock: Optional[bool] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(Product)

    if current_user.role != "admin":
        accessible_warehouse_ids = [w.id for w in current_user.warehouses]
        if accessible_warehouse_ids:
            query = query.filter(
                or_(
                    Product.warehouse_id.in_(accessible_warehouse_ids),
                    Product.warehouse_id.is_(None)
                )
            )

    if search:
        search_term = f"%{search.lower()}%"
        query = query.filter(
            or_(
                func.lower(Product.name).like(search_term),
                func.lower(Product.description).like(search_term)
            )
        )

    if category and category != "Tümü":
        query = query.filter(Product.category == category)

    if status:
        query = query.filter(Product.status == status)

    if warehouse_id:
        query = query.filter(Product.warehouse_id == warehouse_id)

    if low_stock:
        query = query.filter(Product.current_stock <= Product.critical_stock)

    return {"count": query.count()}


@router.get("/stats")
def get_stats(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    query = db.query(Product)
    
    # Filter by user's warehouse access
    if current_user.role != "admin":
        accessible_warehouse_ids = [w.id for w in current_user.warehouses]
        if accessible_warehouse_ids:
            query = query.filter(
                or_(
                    Product.warehouse_id.in_(accessible_warehouse_ids),
                    Product.warehouse_id.is_(None)
                )
            )
        else:
            query = query.filter(Product.warehouse_id.is_(None))

    total = query.count()
    low_stock = query.filter(Product.current_stock <= Product.critical_stock).count()
    broken = query.filter(Product.status == "Bozuk / Kırık").count()
    categories = {}
    for cat in ["Sarf Malzemesi", "Genel / Dayanıklı Malzeme", "Elektronik Bileşen"]:
        categories[cat] = query.filter(Product.category == cat).count()

    return {
        "total_products": total,
        "low_stock_count": low_stock,
        "broken_count": broken,
        "categories": categories
    }


@router.get("/{product_id}", response_model=ProductResponse)
def get_product(product_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    product = db.query(Product).options(
        joinedload(Product.courses),
        joinedload(Product.warehouse),
        joinedload(Product.cabinet),
        joinedload(Product.shelf),
        joinedload(Product.compartment)
    ).filter(Product.id == product_id).first()

    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")

    # Access check
    if current_user.role != "admin":
        accessible_warehouse_ids = [w.id for w in current_user.warehouses]
        if product.warehouse_id is not None and product.warehouse_id not in accessible_warehouse_ids:
            raise HTTPException(status_code=403, detail="Bu ürünü görüntülemek için depoya erişim yetkiniz yok")

    return product_to_response(product)


@router.post("/", response_model=ProductResponse)
def create_product(product_data: ProductCreate, db: Session = Depends(get_db), current_user: User = Depends(get_admin_user)):
    # Check duplicate name in same warehouse
    existing = db.query(Product).filter(
        func.lower(Product.name) == product_data.name.lower(),
        Product.warehouse_id == product_data.warehouse_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Bu depoda aynı isimde bir ürün zaten mevcut")

    product = Product(
        name=product_data.name,
        description=product_data.description,
        image_url=product_data.image_url,
        category=product_data.category,
        status=product_data.status,
        current_stock=product_data.current_stock,
        critical_stock=product_data.critical_stock,
        ideal_stock=product_data.ideal_stock,
        warehouse_id=product_data.warehouse_id,
        cabinet_id=product_data.cabinet_id,
        shelf_id=product_data.shelf_id,
        compartment_id=product_data.compartment_id,
    )

    if product_data.course_ids:
        courses = db.query(Course).filter(Course.id.in_(product_data.course_ids)).all()
        product.courses = courses

    db.add(product)
    db.commit()
    db.refresh(product)

    # Log
    log = StockLog(
        product_id=product.id,
        user_id=current_user.id,
        action="create",
        new_value=str(product.current_stock),
        note=f"Ürün oluşturuldu: {product.name}"
    )
    db.add(log)
    db.commit()

    return product_to_response(product)


@router.put("/{product_id}", response_model=ProductResponse)
def update_product(product_id: int, product_data: ProductUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_admin_user)):
    product = db.query(Product).options(
        joinedload(Product.courses),
        joinedload(Product.warehouse),
        joinedload(Product.cabinet),
        joinedload(Product.shelf),
        joinedload(Product.compartment)
    ).filter(Product.id == product_id).first()

    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")

    # Check name duplicate if name is being changed
    if product_data.name and product_data.name.lower() != product.name.lower():
        wh_id = product_data.warehouse_id if product_data.warehouse_id is not None else product.warehouse_id
        existing = db.query(Product).filter(
            func.lower(Product.name) == product_data.name.lower(),
            Product.warehouse_id == wh_id,
            Product.id != product_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="Bu depoda aynı isimde bir ürün zaten mevcut")

    update_fields = product_data.dict(exclude_unset=True, exclude={"course_ids"})
    for field, value in update_fields.items():
        setattr(product, field, value)

    if product_data.course_ids is not None:
        courses = db.query(Course).filter(Course.id.in_(product_data.course_ids)).all()
        product.courses = courses

    db.commit()
    db.refresh(product)
    return product_to_response(product)

@router.post("/import/excel")
async def import_excel(
    file: UploadFile = File(...),
    warehouse_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_admin_user)
):
    import openpyxl
    import io
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Lütfen geçerli bir Excel dosyası yükleyin (.xlsx veya .xls)")
        
    contents = await file.read()
    
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel dosyası okunamadı: {str(e)}")
        
    # Tüm sayfaları kontrol et ve doğru başlığı bul
    ws = None
    header_row = None
    header_row_idx = 1
    name_idx = -1
    
    for sheet in wb.worksheets:
        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True), 1):
            if not row:
                continue
            headers = [str(h).lower().strip() if h else "" for h in row]
            for i, h in enumerate(headers):
                if h in ["malzeme adı", "ürün adı", "ürün", "malzeme", "ad"] or "malzeme adı" in h or "ürün adı" in h:
                    name_idx = i
                    header_row = headers
                    header_row_idx = idx
                    ws = sheet
                    break
            if header_row:
                break
        if ws:
            break
            
    if not ws or not header_row or name_idx == -1:
        raise HTTPException(status_code=400, detail="Excel dosyasındaki hiçbir sayfada 'Malzeme Adı' veya 'Ürün Adı' içeren bir başlık satırı bulunamadı.")
        
    added_count = 0
    skipped_count = 0
    
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or name_idx >= len(row) or not row[name_idx]:
            continue
            
        name = str(row[name_idx]).strip()
        if not name or name.lower() == "none" or name.lower() == "nan":
            continue
        
        # Kopya kontrolü
        existing = db.query(Product).filter(
            func.lower(Product.name) == name.lower(),
            Product.warehouse_id == warehouse_id
        ).first()
        
        if existing:
            skipped_count += 1
            continue
            
        category = "Genel / Dayanıklı Malzeme"
        description = ""
        current_stock = 0
        critical_stock = 5
        ideal_stock = 100
        status = "Çalışan"
        
        # Diğer sütunları eşleştir
        for i, h in enumerate(headers):
            if i >= len(row) or row[i] is None:
                continue
            val = row[i]
            if "kategori" in h or "category" in h:
                category = str(val).strip()
            elif "açıklama" in h or "desc" in h:
                description = str(val).strip()
            elif "mevcut" in h or ("stok" in h and "kritik" not in h and "ideal" not in h):
                try: current_stock = int(float(val))
                except: pass
            elif "kritik" in h:
                try: critical_stock = int(float(val))
                except: pass
            elif "ideal" in h:
                try: ideal_stock = int(float(val))
                except: pass
            elif "durum" in h or "status" in h:
                status = str(val).strip()
                
        product = Product(
            name=name,
            description=description,
            category=category,
            current_stock=current_stock,
            critical_stock=critical_stock,
            ideal_stock=ideal_stock,
            status=status,
            warehouse_id=warehouse_id
        )
        
        db.add(product)
        added_count += 1
        
    if added_count > 0:
        db.commit()
        
    return {
        "message": "İçe aktarma tamamlandı.",
        "added": added_count,
        "skipped": skipped_count
    }


@router.delete("/{product_id}")
def delete_product(product_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_admin_user)):
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")

    db.delete(product)
    db.commit()
    return {"message": "Ürün silindi"}


@router.post("/{product_id}/stock")
def update_stock(product_id: int, stock_data: StockUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")

    # Check warehouse access
    if current_user.role != "admin":
        accessible_warehouse_ids = [w.id for w in current_user.warehouses]
        if product.warehouse_id is not None and product.warehouse_id not in accessible_warehouse_ids:
            raise HTTPException(status_code=403, detail="Bu deponun stoklarını güncellemeye yetkiniz yok")

    previous = product.current_stock

    if stock_data.action == "add":
        product.current_stock += stock_data.quantity
    elif stock_data.action == "remove":
        product.current_stock = max(0, product.current_stock - stock_data.quantity)
    elif stock_data.action == "set":
        product.current_stock = stock_data.quantity

    log = StockLog(
        product_id=product.id,
        user_id=current_user.id,
        action=stock_data.action,
        quantity_change=stock_data.quantity if stock_data.action != "set" else stock_data.quantity - previous,
        previous_value=str(previous),
        new_value=str(product.current_stock),
        note=stock_data.note
    )
    db.add(log)
    db.commit()

    return {
        "current_stock": product.current_stock,
        "is_critical": product.current_stock <= product.critical_stock
    }


@router.get("/{product_id}/logs", response_model=List)
def get_product_logs(product_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    logs = db.query(StockLog).filter(
        StockLog.product_id == product_id
    ).order_by(StockLog.created_at.desc()).limit(50).all()

    result = []
    for log in logs:
        user_name = None
        if log.user:
            user_name = log.user.full_name
        result.append({
            "id": log.id,
            "product_id": log.product_id,
            "user_id": log.user_id,
            "action": log.action,
            "quantity_change": log.quantity_change,
            "previous_value": log.previous_value,
            "new_value": log.new_value,
            "note": log.note,
            "created_at": log.created_at.replace(tzinfo=timezone.utc).isoformat() if log.created_at else None,
            "user_name": user_name
        })

    return result


@router.post("/upload-image")
async def upload_image(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    # Validate file type
    ext = os.path.splitext(file.filename)[1]
    if ext.lower() not in [".png", ".jpg", ".jpeg", ".gif", ".webp"]:
        raise HTTPException(status_code=400, detail="Geçersiz dosya formatı")

    # Validate file size (5MB max)
    file.file.seek(0, 2)  # Seek to end
    file_size = file.file.tell()
    file.file.seek(0)  # Seek back to start
    
    MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB
    if file_size > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="Dosya boyutu 5MB'dan büyük olamaz")

    # Try Supabase Storage first
    try:
        # Read file content for supabase upload
        file_content = await file.read()
        await file.seek(0) # Reset pointer for potential local fallback
        
        supabase_url = await upload_file_to_supabase(
            file_content, 
            file.filename, 
            file.content_type
        )
        
        if supabase_url:
            return {"url": supabase_url}
    except Exception as e:
        print(f"Supabase upload failed, falling back to local: {e}")

    # Fallback to Local Storage
    filename = f"{uuid.uuid4()}{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)

    with open(filepath, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    return {"url": f"/uploads/{filename}"}
